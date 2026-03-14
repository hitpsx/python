import os
import sys
from openpyxl import Workbook, load_workbook
from collections import defaultdict

def merge_excel_files(input_files, output_file):
    """
    合并多个Excel文件，处理每个文件中的多个sheet
    :param input_files: 输入Excel文件列表
    :param output_file: 输出Excel文件路径
    """
    output_wb = Workbook()
    output_wb.remove(output_wb.active)
    
    sheet_data = defaultdict(list)
    
    for file_path in input_files:
        try:
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = []
                data_rows = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        headers = list(row)
                    else:
                        data_rows.append(list(row))
                
                if headers:
                    sheet_data[sheet_name].append((headers, data_rows))
            wb.close()
        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {e}")
    
    for sheet_name, sheet_items in sheet_data.items():
        output_ws = output_wb.create_sheet(title=sheet_name)
        
        if not sheet_items:
            continue
        
        headers = sheet_items[0][0]
        output_ws.append(headers)
        
        for headers, data_rows in sheet_items:
            for row in data_rows:
                output_ws.append(row)
    
    output_wb.save(output_file)
    print(f"合并完成，结果保存在: {output_file}")

def split_excel_by_columns(input_file, output_dir, split_columns):
    """
    按指定列拆分Excel文件
    :param input_file: 输入Excel文件路径
    :param output_dir: 输出目录
    :param split_columns: 用于拆分的列索引列表（从0开始）
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        wb = load_workbook(input_file)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            data_dict = defaultdict(list)
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = list(row)
                else:
                    key = tuple(row[col] for col in split_columns)
                    data_dict[key].append(list(row))
            
            for key, rows in data_dict.items():
                output_wb = Workbook()
                output_ws = output_wb.active
                output_ws.title = sheet_name
                
                output_ws.append(headers)
                for row in rows:
                    output_ws.append(row)
                
                key_str = "_".join(str(k) for k in key)
                output_file = os.path.join(output_dir, f"{sheet_name}_{key_str}.xlsx")
                output_wb.save(output_file)
                print(f"生成文件: {output_file}")
        wb.close()
    except Exception as e:
        print(f"拆分文件时出错: {e}")

def main():
    if len(sys.argv) < 2:
        print("用法:")
        print("  合并Excel: python excel_tool.py merge <输出文件> <输入文件1> <输入文件2> ...")
        print("  拆分Excel: python excel_tool.py split <输入文件> <输出目录> <列索引1> <列索引2> ...")
        return
    
    command = sys.argv[1]
    
    if command == "merge":
        if len(sys.argv) < 4:
            print("合并命令需要至少一个输入文件")
            return
        output_file = sys.argv[2]
        input_files = sys.argv[3:]
        merge_excel_files(input_files, output_file)
    
    elif command == "split":
        if len(sys.argv) < 5:
            print("拆分命令需要输入文件、输出目录和至少一个列索引")
            return
        input_file = sys.argv[2]
        output_dir = sys.argv[3]
        split_columns = [int(col) for col in sys.argv[4:]]
        split_excel_by_columns(input_file, output_dir, split_columns)
    
    else:
        print("未知命令")

if __name__ == "__main__":
    main()
